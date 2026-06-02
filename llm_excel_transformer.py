"""Standalone LLM-based transformer.

Input:
    A JSON file produced by the PDF parser pipeline, such as `input.output.json`.

Output:
    An Excel workbook that presents the extracted repair items in a tabular form.

This script is intentionally independent from the PDF parser package.
It only consumes the previously generated JSON and produces `.xlsx`.

DeepSeek configuration is reserved at the top of the file.
Do not hardcode secrets in source control; paste your key locally before running.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# DeepSeek configuration
# ---------------------------------------------------------------------------

# Paste your actual API key here before running locally.
# Example format: sk-xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
DEEPSEEK_API_KEY = ""

# DeepSeek official base URL.
DEEPSEEK_BASE_URL = "https://api.deepseek.com"

# Required model.
DEEPSEEK_MODEL = "deepseek-v4-flash"

# Request controls.
DEEPSEEK_TEMPERATURE = 0.0
DEEPSEEK_MAX_TOKENS = 4096
DEEPSEEK_TIMEOUT_SECONDS = 120
DEEPSEEK_RETRY_COUNT = 3
DEFAULT_BATCH_CHARS = 8000


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class SourceItem:
    item_id: str
    title: str
    text: str
    page_refs: list[int]
    source_type: str


@dataclass
class ExcelRow:
    item_no: str
    description: str
    source_id: str = ""
    source_pages: str = ""
    confidence: float = 0.0
    notes: str = ""


# ---------------------------------------------------------------------------
# JSON ingestion
# ---------------------------------------------------------------------------


def load_source_json(json_path: Path) -> dict[str, Any]:
    with json_path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def build_source_items(payload: dict[str, Any]) -> list[SourceItem]:
    """Flatten parser output into compact items for LLM consumption.

    Prefers `sections`, then falls back to `chunks`, then page text blocks.
    """
    sections = payload.get("sections") or []
    chunks = payload.get("chunks") or []
    pages = payload.get("pages") or []

    source_items: list[SourceItem] = []

    # Priority 0: if parser attached explicit table objects, try to extract
    # rows from tables that look like quotation / JOB ITEMS tables. This
    # helps the transformer focus on the actual quotation table rows.
    table_rows: list[SourceItem] = []
    sfi_re = re.compile(r"\d{2,3}\.\d{2,3}(?:\.\d{2})?")
    for page in pages:
        for table in (page.get("tables") or []):
            text = (table.get("text") or "").strip()
            if not text:
                continue
            # crude heuristic: look for JOB ITEMS header, SFI column or 'Item Name' markers
            if any(keyword in text for keyword in ("JOB ITEMS", "Item Name", "SFI", "quotation", "Quotation")):
                # parse lines into candidate rows
                lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
                i = 0
                while i < len(lines):
                    m = sfi_re.search(lines[i])
                    if m:
                        item_no = m.group(0)
                        # next non-code line is likely the item name
                        name = ""
                        desc_lines: list[str] = []
                        if i + 1 < len(lines) and not sfi_re.search(lines[i + 1]):
                            name = lines[i + 1]
                            j = i + 2
                        else:
                            j = i + 1
                        # gather following description lines until next SFI-like code
                        while j < len(lines) and not sfi_re.search(lines[j]):
                            desc_lines.append(lines[j])
                            j += 1
                        desc = "\n".join(desc_lines).strip()
                        table_rows.append(
                            SourceItem(
                                item_id=item_no,
                                title=name,
                                text=desc,
                                page_refs=_safe_int_list([page.get("page_number")]),
                                source_type="table_row",
                            )
                        )
                        i = j
                        continue
                    i += 1
    if table_rows:
        return table_rows

    if sections:
        for index, section in enumerate(sections, start=1):
            source_items.append(
                SourceItem(
                    item_id=str(section.get("section_id") or f"sec_{index:04d}"),
                    title=str(section.get("title") or ""),
                    text=str(section.get("text") or ""),
                    page_refs=[int(p) for p in section.get("page_refs") or ([section.get("page_number")] if section.get("page_number") else []) if p],
                    source_type="section",
                )
            )
        return source_items

    if chunks:
        for index, chunk in enumerate(chunks, start=1):
            source_items.append(
                SourceItem(
                    item_id=str(chunk.get("chunk_id") or f"chunk_{index:04d}"),
                    title=str(chunk.get("context_hint") or ""),
                    text=str(chunk.get("text") or ""),
                    page_refs=_safe_int_list([chunk.get("page_start"), chunk.get("page_end")]),
                    source_type="chunk",
                )
            )
        return source_items

    for page in pages:
        blocks = page.get("blocks") or []
        page_number = page.get("page_number")
        for index, block in enumerate(blocks, start=1):
            if not block.get("text"):
                continue
            source_items.append(
                SourceItem(
                    item_id=f"page_{page_number}_block_{index}",
                    title=str(block.get("block_type") or ""),
                    text=str(block.get("text") or ""),
                    page_refs=_safe_int_list([page_number]),
                    source_type=str(block.get("block_type") or "block"),
                )
            )

    return source_items


def _safe_int_list(values: Iterable[Any]) -> list[int]:
    result: list[int] = []
    for value in values:
        try:
            if value is None:
                continue
            result.append(int(value))
        except (TypeError, ValueError):
            continue
    return result


# ---------------------------------------------------------------------------
# Prompting
# ---------------------------------------------------------------------------


def build_prompt(batch: list[SourceItem]) -> str:
    """Build a compact extraction prompt for one batch."""
    # Minimal schema: prefer these two fields in this order. Traceability
    # fields are allowed but the model must return rows where the
    # primary columns are `item_no` and `description` (description in
    # plain text, suitable for Excel column 2).
    schema_hint = {
        "rows": [
            {
                "item_no": "01.01",
                "description": "Pilot",
                "source_id": "sec_0001",
                "source_pages": [1, 2],
                "confidence": 0.93,
                "notes": "Optional short note when the source is ambiguous."
            }
        ]
    }

    source_lines = []
    for item in batch:
        source_lines.append(
            f"[ID={item.item_id}] [TYPE={item.source_type}] [PAGES={','.join(map(str, item.page_refs)) or '-'}]\n"
            f"TITLE: {item.title}\n"
            f"TEXT:\n{item.text.strip()}\n"
        )

    instruction = (
        "You are converting a ship repair specification into a clean Excel-ready table. "
        "Priority: if the input batch contains rows extracted from a PDF table (source_type == 'table_row'), "
        "use those rows as the primary source. These represent the 'quotation' or 'JOB ITEMS' tables. "
        "Extract logical repair items only and preserve item numbers exactly as they appear (e.g. 140.01). "
        "Output one JSON array `rows` where each element is an object with at least `item_no` and `description`. "
        "The `description` must be the job description and should be placed in the second Excel column. "
        "Do not invent values, do not output extra columns that will shift Excel columns, and do not output markdown or code fences. "
        "If a source row has multiple sub-lines, concatenate them into a single `description` string. "
        "If you cannot determine an item number, leave `item_no` empty but still provide `description`. "
        "Return strict JSON only, no markdown, no code fences."
    )

    return (
        f"{instruction}\n\n"
        f"Required JSON schema example:\n{json.dumps(schema_hint, ensure_ascii=False, indent=2)}\n\n"
        f"Source items:\n\n" + "\n---\n".join(source_lines)
    )


# ---------------------------------------------------------------------------
# DeepSeek client
# ---------------------------------------------------------------------------


def call_deepseek(prompt: str) -> dict[str, Any]:
    if not DEEPSEEK_API_KEY:
        raise RuntimeError(
            "DEEPSEEK_API_KEY is empty. Paste your key at the top of llm_excel_transformer.py before running."
        )

    url = f"{DEEPSEEK_BASE_URL.rstrip('/')}/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
        "Content-Type": "application/json",
    }
    payload = {
        "model": DEEPSEEK_MODEL,
        "temperature": DEEPSEEK_TEMPERATURE,
        "max_tokens": DEEPSEEK_MAX_TOKENS,
        "response_format": {"type": "json_object"},
        "messages": [
            {"role": "system", "content": "Return only valid JSON."},
            {"role": "user", "content": prompt},
        ],
    }

    last_error: Exception | None = None
    for attempt in range(1, DEEPSEEK_RETRY_COUNT + 1):
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=DEEPSEEK_TIMEOUT_SECONDS)
            response.raise_for_status()
            data = response.json()
            content = data["choices"][0]["message"]["content"]
            return extract_json_object(content)
        except Exception as exc:  # noqa: BLE001 - intentional retry wrapper
            last_error = exc
            if attempt < DEEPSEEK_RETRY_COUNT:
                time.sleep(1.5 * attempt)
                continue
            raise RuntimeError(f"DeepSeek request failed after {DEEPSEEK_RETRY_COUNT} attempts: {exc}") from exc

    raise RuntimeError(f"DeepSeek request failed: {last_error}")


def extract_json_object(text: str) -> dict[str, Any]:
    """Extract a top-level JSON object from model output."""
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s*```$", "", text)

    # direct parse first
    try:
        loaded = json.loads(text)
        if isinstance(loaded, dict):
            return loaded
    except json.JSONDecodeError:
        pass

    # fallback: locate first balanced object-ish region
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        candidate = text[start : end + 1]
        loaded = json.loads(candidate)
        if isinstance(loaded, dict):
            return loaded

    raise ValueError("Model response did not contain a valid JSON object.")


# ---------------------------------------------------------------------------
# Batching and row normalization
# ---------------------------------------------------------------------------


def split_batches(items: list[SourceItem], max_chars: int = DEFAULT_BATCH_CHARS) -> list[list[SourceItem]]:
    batches: list[list[SourceItem]] = []
    current: list[SourceItem] = []
    current_len = 0
    for item in items:
        estimated = len(item.title) + len(item.text) + 128
        if current and current_len + estimated > max_chars:
            batches.append(current)
            current = []
            current_len = 0
        current.append(item)
        current_len += estimated
    if current:
        batches.append(current)
    return batches


def normalize_rows(batch_result: dict[str, Any], fallback_source: list[SourceItem]) -> list[ExcelRow]:
    raw_rows = batch_result.get("rows") or batch_result.get("items") or []
    rows: list[ExcelRow] = []
    fallback_by_id = {item.item_id: item for item in fallback_source}

    for index, raw in enumerate(raw_rows, start=1):
        if not isinstance(raw, dict):
            continue
        source_id = str(raw.get("source_id") or raw.get("item_id") or "")
        source = fallback_by_id.get(source_id)
        pages = raw.get("source_pages")
        if not pages and source:
            pages = source.page_refs
        rows.append(
            ExcelRow(
                item_no=str(raw.get("item_no") or raw.get("item") or source_id or f"row_{index:04d}"),
                description=str(raw.get("description") or raw.get("job_description") or raw.get("title") or "").strip(),
                source_id=source_id,
                source_pages=", ".join(map(str, _safe_int_list(pages or []))),
                confidence=_safe_float(raw.get("confidence"), default=0.0),
                notes=str(raw.get("notes") or "").strip(),
            )
        )

    return dedupe_rows(rows)


def dedupe_rows(rows: list[ExcelRow]) -> list[ExcelRow]:
    seen: set[tuple[str, str]] = set()
    unique_rows: list[ExcelRow] = []
    for row in rows:
        key = (row.item_no.strip(), re.sub(r"\s+", " ", row.description).strip().lower())
        if key in seen:
            continue
        seen.add(key)
        unique_rows.append(row)
    return unique_rows


def remap_and_enrich_rows(rows: list[ExcelRow], sources: list[SourceItem]) -> list[ExcelRow]:
    """Remap original item numbers into a compact sequence starting at 01.

    - Main groups are assigned in order of first appearance and numbered 01,02,...
    - Sub-items become XX.YY where YY is the sub-index per main group.
    - Clean obvious prompt artifacts and keep the description concise.
    - If a description is missing, fall back to a short source title or source text excerpt.
    """
    main_map: dict[str, int] = {}
    main_counters: dict[str, int] = {}
    next_main = 1
    src_by_id = {s.item_id: s for s in sources}

    out: list[ExcelRow] = []
    for row in rows:
        orig = (row.item_no or "").strip()
        m = re.match(r"(?P<main>\d+)(?:[.\-](?P<sub>[\d.]+))?", orig)
        if m:
            main_key = m.group("main")
            sub_part = m.group("sub")
        else:
            main_key = row.source_id.split("_")[0] if row.source_id else "0"
            sub_part = None

        if main_key not in main_map:
            main_map[main_key] = next_main
            next_main += 1
            main_counters[main_key] = 0

        if sub_part:
            try:
                sub_num = int(str(sub_part).split(".")[-1])
            except Exception:
                main_counters[main_key] += 1
                sub_num = main_counters[main_key]
            main_counters[main_key] = max(main_counters[main_key], sub_num)
        else:
            main_counters[main_key] += 1
            sub_num = main_counters[main_key]

        new_main = main_map[main_key]
        new_item_no = f"{new_main:02d}.{sub_num:02d}"

        desc = _clean_description_text(row.description or "")
        if not desc:
            src = src_by_id.get(row.source_id)
            if src:
                fallback = _clean_description_text(src.title or "")
                if not fallback:
                    fallback = _first_meaningful_line(src.text or "")
                desc = fallback

        out.append(
            ExcelRow(
                item_no=new_item_no,
                description=desc,
                source_id=row.source_id,
                source_pages=row.source_pages,
                confidence=row.confidence,
                notes=row.notes,
            )
        )

    return out


def _first_meaningful_line(text: str) -> str:
    for line in (text or "").splitlines():
        line = line.strip()
        if not line:
            continue
        if re.fullmatch(r"(?i)(source context:.*|description:|job items|sfi\*?|item name|priority|type|cost type|vessel|location|type/serial no|component|start date|end date|status|responsible|maker|yard)", line):
            continue
        return line
    return ""


def _clean_description_text(text: str) -> str:
    """Remove prompt artifacts and collapse a description to a readable single line."""
    if not text:
        return ""

    lines: list[str] = []
    seen_norm_lines: set[str] = set()
    for raw_line in str(text).replace("\r", "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        if re.fullmatch(r"(?i)source context:.*", line):
            continue
        if re.fullmatch(r"(?i)(description:|job items|sfi\*?|item name|priority|type|cost type|vessel|location|type/serial no|component|start date|end date|status|responsible|maker|yard)", line):
            continue
        normalized_line = re.sub(r"\s+", " ", line).strip().rstrip(".").lower()
        if normalized_line in seen_norm_lines:
            continue
        seen_norm_lines.add(normalized_line)
        lines.append(line)

    if not lines:
        return ""

    cleaned = " ".join(lines)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = re.sub(r"(?i)^source context:\s*", "", cleaned).strip()
    return cleaned


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------


def write_excel(rows: list[ExcelRow], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Repair Items"

    headers = ["Item No.", "Description of job"]
    ws.append(headers)
    ws.freeze_panes = "A2"

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([row.item_no, row.description])

    # a compact, readable layout similar to the screenshot.
    widths = {
        1: 14,
        2: 96,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in ws.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[0].alignment = Alignment(vertical="top")

    trace = wb.create_sheet("Traceability")
    trace.sheet_state = "hidden"
    trace_headers = ["Item No.", "Source ID", "Source Pages", "Confidence", "Notes"]
    trace.append(trace_headers)
    for cell in trace[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        trace.append([
            row.item_no,
            row.source_id,
            row.source_pages,
            row.confidence,
            row.notes,
        ])

    trace_widths = {1: 14, 2: 26, 3: 18, 4: 12, 5: 40}
    for col_idx, width in trace_widths.items():
        trace.column_dimensions[get_column_letter(col_idx)].width = width

    for row in trace.iter_rows(min_row=2):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        for cell in row:
            if cell.column != 5:
                cell.alignment = Alignment(vertical="top")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------


def transform(json_path: Path, output_xlsx: Path, batch_chars: int = DEFAULT_BATCH_CHARS) -> list[ExcelRow]:
    payload = load_source_json(json_path)
    source_items = build_source_items(payload)
    if not source_items:
        raise RuntimeError("No source items found in the input JSON.")

    all_rows: list[ExcelRow] = []
    batches = split_batches(source_items, max_chars=batch_chars)

    for batch_index, batch in enumerate(batches, start=1):
        prompt = build_prompt(batch)
        result = call_deepseek(prompt)
        rows = normalize_rows(result, batch)
        all_rows.extend(rows)

    all_rows = dedupe_rows(all_rows)
    # remap to compact numbering starting at 01 and enrich short descriptions
    final_rows = remap_and_enrich_rows(all_rows, source_items)
    write_excel(final_rows, output_xlsx)
    return final_rows


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Convert parser JSON into an Excel workbook using DeepSeek.")
    parser.add_argument("input_json", help="Path to the parser output JSON, e.g. input.output.json")
    parser.add_argument("output_xlsx", help="Path to the Excel workbook to create")
    parser.add_argument("--batch-chars", type=int, default=DEFAULT_BATCH_CHARS, help="Approximate character budget per LLM batch")
    parser.add_argument("--dry-run", action="store_true", help="Print the planned batches without calling the API")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    input_json = Path(args.input_json)
    output_xlsx = Path(args.output_xlsx)

    payload = load_source_json(input_json)
    source_items = build_source_items(payload)
    batches = split_batches(source_items, max_chars=args.batch_chars)

    if args.dry_run:
        print(f"Input: {input_json}")
        print(f"Output: {output_xlsx}")
        print(f"Source items: {len(source_items)}")
        print(f"Batches: {len(batches)}")
        for idx, batch in enumerate(batches[:5], start=1):
            print(f"  Batch {idx}: {len(batch)} items, first ids: {[item.item_id for item in batch[:3]]}")
        return 0

    rows = transform(input_json, output_xlsx, batch_chars=args.batch_chars)
    print(f"Wrote {len(rows)} rows to {output_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
